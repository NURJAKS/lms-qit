from typing import List, Any, Optional, Iterable
import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def _fill_sheet(ws, headers: List[str] | None, rows: Iterable[List[Any]]) -> None:
    if headers:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width


def generate_multi_sheet_xlsx_response(
    sheets: list[tuple[str, List[str] | None, Iterable[List[Any]]]],
    filename: str,
) -> StreamingResponse:
    """
    Build an XLSX with multiple sheets. Sheet titles are truncated to 31 chars (Excel limit).
    """
    wb = Workbook()
    wb.remove(wb.active)
    for raw_title, headers, rows in sheets:
        title = (raw_title or "Sheet")[:31]
        ws = wb.create_sheet(title=title)
        _fill_sheet(ws, headers, rows)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )


def generate_xlsx_response(
    data: Iterable[List[Any]], 
    filename: str, 
    headers: List[str] | None = None,
    sheet_name: str = "Data"
) -> StreamingResponse:
    """
    Generates a StreamingResponse for an XLSX file.
    Applies professional styling: bold headers and auto-adjusted column widths.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    _fill_sheet(ws, headers, data)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
    )
